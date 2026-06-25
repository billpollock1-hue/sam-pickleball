import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = Path("data/master_history_raw.csv")
OUTPUT_FILE = Path("output/pickleball_2026_summary_report.xlsx")
YEAR = 2026

OUTPUT_FILE.parent.mkdir(exist_ok=True)

def split_team(team):
    return [p.strip() for p in str(team).split("/") if p.strip()]

df = pd.read_csv(INPUT_FILE)
df.columns = [c.strip() for c in df.columns]

df["posted_dt"] = pd.to_datetime(df["posted"], errors="coerce")
df = df[df["posted_dt"].dt.year == YEAR].copy()
df = df.reset_index(drop=True)
df["Match_Order"] = df.index + 1

df["winning_score"] = pd.to_numeric(df["winning_score"], errors="coerce")
df["losing_score"] = pd.to_numeric(df["losing_score"], errors="coerce")
df["point_diff"] = df["winning_score"] - df["losing_score"]
df["weekday"] = df["posted_dt"].dt.day_name()
df["scoreline"] = df["winning_score"].astype("Int64").astype(str) + "-" + df["losing_score"].astype("Int64").astype(str)

player_rows = []

for _, r in df.iterrows():
    winners = split_team(r["winning_team"])
    losers = split_team(r["losing_team"])

    for p in winners:
        player_rows.append({
            "Player": p,
            "Date": r["posted_dt"].date(),
            "Weekday": r["weekday"],
            "Match_Order": r["Match_Order"],
            "Result": "Win",
            "Points For": r["winning_score"],
            "Points Against": r["losing_score"],
            "Point Diff": r["point_diff"],
            "Scoreline": r["scoreline"],
            "Pool": r.get("pool", "")
        })

    for p in losers:
        player_rows.append({
            "Player": p,
            "Date": r["posted_dt"].date(),
            "Weekday": r["weekday"],
            "Match_Order": r["Match_Order"],
            "Result": "Loss",
            "Points For": r["losing_score"],
            "Points Against": r["winning_score"],
            "Point Diff": -r["point_diff"],
            "Scoreline": r["scoreline"],
            "Pool": r.get("pool", "")
        })

player_log = pd.DataFrame(player_rows)

games_by_player = (
    player_log
    .groupby("Player")
    .agg(
        Games=("Player", "count"),
        Wins=("Result", lambda x: (x == "Win").sum()),
        Losses=("Result", lambda x: (x == "Loss").sum()),
        Avg_Point_Diff=("Point Diff", "mean"),
        Points_For=("Points For", "sum"),
        Points_Against=("Points Against", "sum")
    )
    .reset_index()
)

games_by_player["Avg_Point_Diff"] = games_by_player["Avg_Point_Diff"].round(1)
games_by_player["Win %"] = games_by_player["Wins"] / games_by_player["Games"]
games_by_player = games_by_player.sort_values(["Games", "Player"], ascending=[False, True])

weekday_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

games_by_weekday = (
    df.groupby("weekday")
    .size()
    .reindex(weekday_order, fill_value=0)
    .reset_index(name="Games")
    .rename(columns={"weekday": "Weekday"})
)

games_by_player_weekday = (
    player_log
    .pivot_table(index="Player", columns="Weekday", values="Result", aggfunc="count", fill_value=0)
    .reindex(columns=weekday_order, fill_value=0)
    .reset_index()
)
games_by_player_weekday["Total Games"] = games_by_player_weekday[weekday_order].sum(axis=1)
games_by_player_weekday = games_by_player_weekday.sort_values("Total Games", ascending=False)

score_distribution = (
    df.groupby("scoreline")
    .size()
    .reset_index(name="Games")
    .sort_values("Games", ascending=False)
)

score_distribution_by_player = (
    player_log
    .pivot_table(index="Player", columns="Scoreline", values="Result", aggfunc="count", fill_value=0)
    .reset_index()
)

point_diff_distribution = (
    df.groupby("point_diff")
    .size()
    .reset_index(name="Games")
    .sort_values("point_diff")
)

point_diff_by_player = (
    player_log
    .pivot_table(index="Player", columns="Point Diff", values="Result", aggfunc="count", fill_value=0)
    .reset_index()
)

games_by_date = (
    df.groupby(df["posted_dt"].dt.date)
    .size()
    .reset_index(name="Games")
    .rename(columns={"posted_dt": "Date"})
)

# Game-number performance:
# Only include player-days where the player played exactly 6 games.
# Exclude partial days, but count excluded player-days and excluded individual games.
player_day_counts = (
    player_log
    .groupby(["Player", "Date"])
    .size()
    .reset_index(name="Games_Played_That_Day")
)

complete_player_days = player_day_counts[player_day_counts["Games_Played_That_Day"] == 6].copy()
excluded_player_days = player_day_counts[player_day_counts["Games_Played_That_Day"] != 6].copy()

eligible_log = player_log.merge(
    complete_player_days[["Player", "Date"]],
    on=["Player", "Date"],
    how="inner"
).copy()

eligible_log = eligible_log.sort_values(["Player", "Date", "Match_Order"])
eligible_log["Game_Number"] = eligible_log.groupby(["Player", "Date"]).cumcount() + 1

game_number_performance = (
    eligible_log
    .groupby(["Player", "Game_Number"])
    .agg(
        Games_Included=("Result", "count"),
        Wins=("Result", lambda x: (x == "Win").sum()),
        Losses=("Result", lambda x: (x == "Loss").sum()),
        Avg_Point_Diff=("Point Diff", "mean"),
        Points_For=("Points For", "sum"),
        Points_Against=("Points Against", "sum")
    )
    .reset_index()
)

game_number_performance["Win %"] = game_number_performance["Wins"] / game_number_performance["Games_Included"]
game_number_performance["Avg_Point_Diff"] = game_number_performance["Avg_Point_Diff"].round(1)

excluded_by_player = (
    excluded_player_days
    .groupby("Player")
    .agg(
        Excluded_Player_Days=("Date", "count"),
        Excluded_Games=("Games_Played_That_Day", "sum")
    )
    .reset_index()
)

game_number_performance = game_number_performance.merge(
    excluded_by_player,
    on="Player",
    how="left"
)

game_number_performance["Excluded_Player_Days"] = game_number_performance["Excluded_Player_Days"].fillna(0).astype(int)
game_number_performance["Excluded_Games"] = game_number_performance["Excluded_Games"].fillna(0).astype(int)

game_number_performance = game_number_performance[
    [
        "Player",
        "Game_Number",
        "Games_Included",
        "Wins",
        "Losses",
        "Win %",
        "Avg_Point_Diff",
        "Points_For",
        "Points_Against",
        "Excluded_Player_Days",
        "Excluded_Games"
    ]
].sort_values(["Player", "Game_Number"])

excluded_game_number_days = excluded_player_days.sort_values(["Player", "Date"])

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    games_by_player.to_excel(writer, sheet_name="Games by Player", index=False)
    games_by_weekday.to_excel(writer, sheet_name="Games by Weekday", index=False)
    games_by_player_weekday.to_excel(writer, sheet_name="Player by Weekday", index=False)
    score_distribution.to_excel(writer, sheet_name="Score Distribution", index=False)
    score_distribution_by_player.to_excel(writer, sheet_name="Score Dist by Player", index=False)
    point_diff_distribution.to_excel(writer, sheet_name="Point Diff Distribution", index=False)
    point_diff_by_player.to_excel(writer, sheet_name="Point Diff by Player", index=False)
    games_by_date.to_excel(writer, sheet_name="Games by Date", index=False)
    game_number_performance.to_excel(writer, sheet_name="Game Number Perf", index=False)
    excluded_game_number_days.to_excel(writer, sheet_name="Game Number Exclusions", index=False)
    player_log.to_excel(writer, sheet_name="Player Game Log", index=False)

wb = load_workbook(OUTPUT_FILE)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9D9D9")
header_border = Border(bottom=thin)

def format_sheet(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = header_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)

    if "Player" in [c.value for c in ws[1]]:
        player_col = [c.value for c in ws[1]].index("Player") + 1
        ws.column_dimensions[get_column_letter(player_col)].width = 24

for ws in wb.worksheets:
    format_sheet(ws)

# Specific formatting
ws = wb["Games by Player"]
headers = [c.value for c in ws[1]]
for row in range(2, ws.max_row + 1):
    ws.cell(row, headers.index("Avg_Point_Diff") + 1).number_format = "0.0"
    ws.cell(row, headers.index("Win %") + 1).number_format = "0.0%"

ws = wb["Game Number Perf"]
headers = [c.value for c in ws[1]]
for row in range(2, ws.max_row + 1):
    ws.cell(row, headers.index("Win %") + 1).number_format = "0.0%"
    ws.cell(row, headers.index("Avg_Point_Diff") + 1).number_format = "0.0"

ws = wb["Games by Date"]
ws.column_dimensions["A"].width = 14
for row in range(2, ws.max_row + 1):
    ws.cell(row, 1).number_format = "m/d/yyyy"

ws = wb["Game Number Exclusions"]
ws.column_dimensions["B"].width = 14
for row in range(2, ws.max_row + 1):
    ws.cell(row, 2).number_format = "m/d/yyyy"

wb.save(OUTPUT_FILE)

print(f"Built summary workbook: {OUTPUT_FILE}")

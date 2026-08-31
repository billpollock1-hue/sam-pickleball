#!/usr/bin/env python3
"""
compute_format_tracker_data.py

Reads Player_Game_Log from the live model workbook and computes:

  1. Daily rating series: average player_pre_rating per day + 7-day
     rolling average, for the existing bar/line chart.

  2. Per-player participation: distinct play-days in a fixed 60-day
     window before the format change (2026-06-26 to 2026-08-23) versus
     the post-change window (2026-08-24 onward, growing as more data
     comes in). The pre-window is intentionally fixed, not rolling --
     re-running this script next week compares against the same
     pre-change baseline, only the post side grows. A 60-day pre-window
     (not a longer one, e.g. 182-day leaderboard-eligibility window) was
     chosen deliberately to avoid blending in broader seasonal
     participation shifts unrelated to the format change itself.

     Participation is expressed as a rate (player's distinct play-days
     / distinct play-days available in that window), not a raw count,
     since the post window is currently much shorter than the pre
     window and a raw-count comparison would be misleading until it
     catches up.

Writes launcher/format_tracker_data.json for format_tracker.html to consume.

Self-contained: no dependency on any other pipeline module. Safe to run
standalone at any time; always recomputes from the full history in the
workbook (single authoritative data source, no parallel calculation).
"""

import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

# ---- Config -----------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = REPO_ROOT / "output" / "pickleball_model_latest.xlsx"
OUTPUT_PATH = REPO_ROOT / "launcher" / "format_tracker_data.json"
SHEET_NAME = "Player_Game_Log"

# Format change effective date (shuffle-format change went live this date)
FORMAT_CHANGE_DATE = date(2026, 8, 24)

ROLLING_WINDOW_DAYS = 7

# Fixed 60-day pre-change baseline window for participation comparison.
# Deliberately NOT the 182-day leaderboard-eligibility window -- that
# window is long enough to blend in seasonal participation shifts
# (e.g. snowbird patterns) unrelated to the format change.
PRE_WINDOW_DAYS = 60
PRE_WINDOW_START = FORMAT_CHANGE_DATE - timedelta(days=PRE_WINDOW_DAYS)
PRE_WINDOW_END = FORMAT_CHANGE_DATE - timedelta(days=1)


def load_game_log(workbook_path: Path) -> pd.DataFrame:
    if not workbook_path.exists():
        print(f"ERROR: workbook not found at {workbook_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found in workbook", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    df = pd.DataFrame(rows, columns=header)
    wb.close()

    if "include_in_ratings" in df.columns:
        df = df[df["include_in_ratings"] == "Yes"].copy()

    df["posted_dt"] = pd.to_datetime(df["posted_dt"])
    df["play_date"] = df["posted_dt"].dt.date
    return df


def compute_daily_series(df: pd.DataFrame) -> pd.DataFrame:
    daily = (
        df.groupby("play_date")["player_pre_rating"]
        .agg(avg_rating="mean", count="count")
        .reset_index()
        .sort_values("play_date")
    )

    daily["rolling_avg"] = (
        daily["avg_rating"]
        .rolling(window=ROLLING_WINDOW_DAYS, min_periods=1)
        .mean()
    )

    return daily


def compute_participation(df: pd.DataFrame) -> dict:
    pre_df = df[(df["play_date"] >= PRE_WINDOW_START) & (df["play_date"] <= PRE_WINDOW_END)]
    post_df = df[df["play_date"] >= FORMAT_CHANGE_DATE]

    pre_available_dates = pre_df["play_date"].nunique()
    post_available_dates = post_df["play_date"].nunique()

    pre_by_player = pre_df.groupby("player")["play_date"].nunique()
    post_by_player = post_df.groupby("player")["play_date"].nunique()

    all_players = sorted(set(pre_by_player.index) | set(post_by_player.index))

    players = []
    for p in all_players:
        pre_days = int(pre_by_player.get(p, 0))
        post_days = int(post_by_player.get(p, 0))
        pre_rate = (pre_days / pre_available_dates) if pre_available_dates else 0.0
        post_rate = (post_days / post_available_dates) if post_available_dates else 0.0

        if pre_days > 0 and post_days == 0:
            status = "dropped"
        elif pre_days == 0 and post_days > 0:
            status = "new"
        else:
            status = "continuing"

        players.append({
            "player": p,
            "pre_days": pre_days,
            "post_days": post_days,
            "pre_rate": round(pre_rate, 3),
            "post_rate": round(post_rate, 3),
            "rate_delta": round(post_rate - pre_rate, 3),
            "status": status,
        })

    # Dropped players first (most actionable), then by biggest negative
    # rate change, so the players worth a second look surface at the top.
    status_order = {"dropped": 0, "continuing": 1, "new": 2}
    players.sort(key=lambda r: (status_order[r["status"]], r["rate_delta"]))

    return {
        "pre_window_start": PRE_WINDOW_START.isoformat(),
        "pre_window_end": PRE_WINDOW_END.isoformat(),
        "pre_available_dates": int(pre_available_dates),
        "post_available_dates": int(post_available_dates),
        "players": players,
    }


def build_output(daily: pd.DataFrame, participation: dict) -> dict:
    pre = daily[daily["play_date"] < FORMAT_CHANGE_DATE]
    post = daily[daily["play_date"] >= FORMAT_CHANGE_DATE]

    last_pre_date = pre["play_date"].max() if not pre.empty else None
    first_post_date = post["play_date"].min() if not post.empty else None

    days = []
    for _, row in daily.iterrows():
        days.append({
            "date": row["play_date"].isoformat(),
            "avg_rating": round(float(row["avg_rating"]), 2),
            "rolling_avg": round(float(row["rolling_avg"]), 2),
            "count": int(row["count"]),
            "period": "before" if row["play_date"] < FORMAT_CHANGE_DATE else "after",
        })

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "format_change_date": FORMAT_CHANGE_DATE.isoformat(),
        "last_pre_change_date": last_pre_date.isoformat() if last_pre_date else None,
        "first_post_change_date": first_post_date.isoformat() if first_post_date else None,
        "rolling_window_days": ROLLING_WINDOW_DAYS,
        "days": days,
        "participation": participation,
    }


def main():
    df = load_game_log(WORKBOOK_PATH)
    daily = compute_daily_series(df)
    participation = compute_participation(df)
    output = build_output(daily, participation)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    dropped = sum(1 for p in participation["players"] if p["status"] == "dropped")
    new = sum(1 for p in participation["players"] if p["status"] == "new")

    print(f"Wrote {len(output['days'])} days to {OUTPUT_PATH}")
    print(f"Cutover: last pre-change day = {output['last_pre_change_date']}, "
          f"first post-change day = {output['first_post_change_date']}")
    print(f"Participation: pre-window {participation['pre_window_start']} to "
          f"{participation['pre_window_end']} ({participation['pre_available_dates']} play dates), "
          f"post-window from {FORMAT_CHANGE_DATE.isoformat()} "
          f"({participation['post_available_dates']} play dates so far)")
    print(f"  {len(participation['players'])} total players tracked, "
          f"{dropped} dropped, {new} new")


if __name__ == "__main__":
    main()

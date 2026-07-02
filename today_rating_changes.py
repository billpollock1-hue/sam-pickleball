#!/usr/bin/env python3
"""
Produces an Excel workbook showing game-by-game and summary rating changes
for all players who participated in today's SAM shootout.
"""

import sys
sys.path.insert(0, "/Users/billpollock/Documents/SAM Pickleball/PickleballModel")

import math
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Import engine constants & functions ──────────────────────────────────────
from pickleball_engine_v2 import (
    BASE_ELO, K_FACTOR, PROVISIONAL_K_START, PROVISIONAL_K_GAMES,
    PLACEHOLDERS, GUEST_OUTPUT_EXCLUSIONS, MANUAL_NAME_FIXES,
    norm, split_team, apply_manual_fix, team_has_placeholder,
    margin_multiplier, expected, game_position_decay, _provisional_k,
)

TODAY = pd.Timestamp("2026-07-02").date()
DATA_PATH = Path("/Users/billpollock/Documents/SAM Pickleball/PickleballModel/data/master_history_raw.csv")
OUT_PATH  = Path("/Users/billpollock/Documents/SAM Pickleball/PickleballModel/output/today_rating_changes.xlsx")

# ── Load & prepare raw data ───────────────────────────────────────────────────
raw = pd.read_csv(DATA_PATH)
raw["posted_dt"] = pd.to_datetime(raw["posted"], errors="coerce")
raw = raw.sort_values("posted_dt").reset_index(drop=True)

# Apply name fixes and placeholder / guest exclusion flags
raw["winning_team"] = raw.apply(lambda r: apply_manual_fix(r["winning_team"], r["posted_dt"]), axis=1)
raw["losing_team"]  = raw.apply(lambda r: apply_manual_fix(r["losing_team"],  r["posted_dt"]), axis=1)
raw["exclude_match"] = raw.get("exclude_match", False).fillna(False).astype(bool)
raw["include_in_ratings"] = ~(
    raw["winning_team"].apply(team_has_placeholder) |
    raw["losing_team"].apply(team_has_placeholder)  |
    raw["exclude_match"]
)

# Pre-count total games per player (needed for decay)
player_total_games = defaultdict(int)
for _, r in raw.iterrows():
    if r["include_in_ratings"]:
        for p in split_team(r["winning_team"]) + split_team(r["losing_team"]):
            if p:
                player_total_games[p] += 1

# ── Run full Elo history ──────────────────────────────────────────────────────
ratings              = defaultdict(lambda: BASE_ELO)
player_game_count    = defaultdict(int)
player_games_window  = defaultdict(list)
game_rows            = []

for match_id, (_, r) in enumerate(raw.iterrows(), start=1):
    w1, w2 = split_team(r["winning_team"])
    l1, l2 = split_team(r["losing_team"])
    sw, sl  = int(r["winning_score"]), int(r["losing_score"])
    include = bool(r["include_in_ratings"])
    posted  = r["posted_dt"]

    snap = {p: ratings[p] for p in [w1, w2, l1, l2] if p}
    for p in [w1, w2, l1, l2]:
        if p and p not in snap:
            snap[p] = BASE_ELO

    team_win_pre  = (snap[w1] + snap[w2]) / 2
    team_lose_pre = (snap[l1] + snap[l2]) / 2
    exp_win = 1 / (1 + 10 ** ((team_lose_pre - team_win_pre) / 400))
    mult    = min(math.log(abs(sw - sl) + 1), 2.0)

    if include:
        k_w1 = _provisional_k(player_game_count[w1] + 1)
        k_w2 = _provisional_k(player_game_count[w2] + 1)
        k_l1 = _provisional_k(player_game_count[l1] + 1)
        k_l2 = _provisional_k(player_game_count[l2] + 1)
    else:
        k_w1 = k_w2 = k_l1 = k_l2 = 0.0

    def pos(p):
        n = len(player_games_window[p]) + 1
        return min(n, 60)

    def decay(p):
        return game_position_decay(pos(p), total_games=player_total_games[p]) if include else 1.0

    d_w1 = round(k_w1 * (1 - exp_win) * mult * decay(w1), 2)
    d_w2 = round(k_w2 * (1 - exp_win) * mult * decay(w2), 2)
    d_l1 = round(k_l1 * (0 - (1 - exp_win)) * mult * decay(l1), 2)
    d_l2 = round(k_l2 * (0 - (1 - exp_win)) * mult * decay(l2), 2)

    is_today = posted.date() == TODAY if pd.notna(posted) else False

    if is_today:
        pool = str(r.get("pool", "")).strip()
        for player, partner, opp1, opp2, is_win, pf, pa, delta in [
            (w1, w2, l1, l2, 1, sw, sl, d_w1),
            (w2, w1, l1, l2, 1, sw, sl, d_w2),
            (l1, l2, w1, w2, 0, sl, sw, d_l1),
            (l2, l1, w1, w2, 0, sl, sw, d_l2),
        ]:
            if not player:
                continue
            pre  = snap[player]
            post = round(pre + delta, 2) if include else pre
            team_pre = (snap[player] + snap[partner]) / 2 if partner else snap[player]
            opp_pre  = (snap[opp1]   + snap[opp2])   / 2 if opp2    else snap[opp1]
            game_rows.append({
                "Time":             posted.strftime("%I:%M %p"),
                "Pool":             pool,
                "Player":           player,
                "Partner":          partner,
                "Opponent 1":       opp1,
                "Opponent 2":       opp2,
                "W/L":              "W" if is_win else "L",
                "Score":            f"{pf}–{pa}",
                "Team Rating":      round(team_pre),
                "Opp Rating":       round(opp_pre),
                "Rating Gap":       round(abs(team_pre - opp_pre)),
                "Pre-Game Rating":  round(pre),
                "Rating Change":    round(delta),
                "Post-Game Rating": round(post),
            })

    if include:
        for p, d in [(w1, d_w1), (w2, d_w2), (l1, d_l1), (l2, d_l2)]:
            if p:
                ratings[p] = round(snap[p] + d, 2)
                player_game_count[p] += 1
                player_games_window[p].append(match_id)
                if len(player_games_window[p]) > 60:
                    player_games_window[p] = player_games_window[p][-60:]

if not game_rows:
    print("No games found for today.")
    sys.exit(1)

games_df = pd.DataFrame(game_rows)

# ── Summary per player ────────────────────────────────────────────────────────
summary_rows = []
for player, grp in games_df.groupby("Player"):
    summary_rows.append({
        "Player":           player,
        "Games":            len(grp),
        "Wins":             (grp["W/L"] == "W").sum(),
        "Losses":           (grp["W/L"] == "L").sum(),
        "Start Rating":     grp["Pre-Game Rating"].iloc[0],
        "End Rating":       grp["Post-Game Rating"].iloc[-1],
        "Total Change":     grp["Rating Change"].sum().round(),
        "Avg Opp Rating":   round(grp["Opp Rating"].mean()),
    })

summary_df = (pd.DataFrame(summary_rows)
              .sort_values("Total Change", ascending=False)
              .reset_index(drop=True))

# ── Write Excel ───────────────────────────────────────────────────────────────
wb = Workbook()

# Styles
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
POS_FILL  = PatternFill("solid", fgColor="E2EFDA")
NEG_FILL  = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
CTR       = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left", vertical="center")
THIN      = Side(style="thin", color="CCCCCC")
BDR       = Border(bottom=THIN)

def write_sheet(ws, df, title):
    ws.title = title
    cols = list(df.columns)

    # Header row
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CTR
    ws.row_dimensions[1].height = 18

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        alt = (ri % 2 == 0)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = BDR

            # Alignment: left for text, center for numbers
            c.alignment = LEFT if isinstance(val, str) else CTR

            # Highlight rating change column
            if col in ("Rating Change", "Total Change"):
                if isinstance(val, (int, float)):
                    if val > 0:
                        c.fill = POS_FILL
                    elif val < 0:
                        c.fill = NEG_FILL
                    else:
                        c.fill = ALT_FILL if alt else PatternFill()
            elif alt:
                c.fill = ALT_FILL
        ws.row_dimensions[ri].height = 16

    # Auto column widths
    for ci, col in enumerate(cols, 1):
        max_len = max(len(str(col)), *(len(str(df.iloc[r][col])) for r in range(len(df))))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)

    ws.freeze_panes = "A2"


# Sheet 1: Game by game (sorted by time then pool)
ws1 = wb.active
games_sorted = games_df.sort_values(["Time", "Pool", "Player"])
write_sheet(ws1, games_sorted, "Game by Game")

# Sheet 2: Player summary
ws2 = wb.create_sheet("Player Summary")
write_sheet(ws2, summary_df, "Player Summary")

OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")

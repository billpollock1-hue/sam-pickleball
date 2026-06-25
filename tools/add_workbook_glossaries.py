from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

xlsx = Path("output/pickleball_model_latest.xlsx")
wb = load_workbook(xlsx)

glossaries = {
    "Leaderboard": {
        "Player Rating": "Current displayed rating based on recent rated games, confidence, freshness, and display compression.",
        "Games Used": "Number of recent rated games used in the leaderboard calculation.",
        "Win %": "Actual wins divided by games used.",
        "Expected Win %": "Average expected win probability based on pre-match team ratings.",
        "Win % vs Expected": "Actual Win % minus Expected Win %. Positive means the player outperformed expectations.",
        "Avg Point Diff": "Average points scored minus points allowed.",
        "Avg Matchup Edge": "Average pre-match rating advantage of the player's team versus opponents.",
        "Freshness Tier": "Current, Aging, Stale, or Inactive based on recency of play.",
        "Confidence Tier": "Reliability category based mainly on number of games used.",
    },
    "Results vs Expectation": {
        "Win % vs Expected": "Actual Win % minus Expected Win %. This is the main performance diagnostic.",
        "Avg Matchup Edge": "Positive means easier average matchups; negative means tougher average matchups.",
    },
    "AB Court Planning": {
        "Rule 1": "A/B split using rating threshold of 1000.",
        "Rule 2": "A/B split using rating threshold of 950 with lagged snapshot ratings.",
        "Courts": "Estimated courts needed assuming 4 players per court.",
        "Missing Snapshot Ratings": "Players without a rating in the applicable snapshot.",
    },
    "Session Effects": {
        "G1 Gap": "Actual Win % minus Expected Win % in Game 1.",
        "G2-5 Gap": "Actual Win % minus Expected Win % in Games 2 through 5.",
        "G1 Effect": "G1 Gap minus G2-5 Gap. Negative suggests a slow start.",
        "G6 Effect": "G6 Gap minus G2-5 Gap. Negative suggests late-session fade.",
    },
    "Recent Trends": {
        "Last 15 Gap": "Actual Win % minus Expected Win % over the most recent 15 rated games.",
        "Previous 15 Gap": "Same metric for games 16 through 30 back.",
        "Trend Change": "Last 15 Gap minus Previous 15 Gap.",
        "Last 30 Gap": "Actual Win % minus Expected Win % over the most recent 30 rated games.",
        "Trend Confidence": "High, Medium, or Low based on available games in the comparison windows.",
    },
    "Recent Best Worst Day": {
        "Best Day Gap": "Best single-day Actual Win % minus Expected Win % within the recent window.",
        "Worst Day Gap": "Worst single-day Actual Win % minus Expected Win % within the recent window.",
        "Record": "Player's win-loss record on that date.",
        "Margin": "Total points scored minus points allowed on that date.",
    },
    "Rating History": {
        "Rating values": "Daily end-of-day leaderboard-scale rating. Ratings only change on dates when the player played.",
        "Highlighted cells": "Dates when the player actually played.",
    },
    "Credibility Sensitivity": {
        "CONF": "Sensitivity test showing how rankings change under different confidence-adjustment strengths.",
        "CONF 20": "Original, more conservative credibility adjustment.",
        "CONF 10": "Current preferred credibility adjustment.",
    },
    "Less Active": {
        "Less Active": "Players with enough history but less recent activity than current leaderboard players.",
    },
}

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
sub_fill = PatternFill("solid", fgColor="D9EAF7")
sub_font = Font(color="1F4E78", bold=True)
thin = Side(style="thin", color="D9D9D9")

for sheet, items in glossaries.items():
    if sheet not in wb.sheetnames:
        continue

    ws = wb[sheet]
    start = ws.max_row + 3

    ws.cell(start, 1).value = "Column Glossary"
    ws.cell(start, 1).fill = header_fill
    ws.cell(start, 1).font = header_font
    ws.cell(start, 1).alignment = Alignment(horizontal="left")
    ws.cell(start, 2).fill = header_fill

    ws.cell(start + 1, 1).value = "Term"
    ws.cell(start + 1, 2).value = "Meaning"

    for c in [1, 2]:
        cell = ws.cell(start + 1, c)
        cell.fill = sub_fill
        cell.font = sub_font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    r = start + 2
    for term, meaning in items.items():
        ws.cell(r, 1).value = term
        ws.cell(r, 2).value = meaning
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 1

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 10, 24)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 10, 80)

wb.save(xlsx)
print(f"Added workbook glossaries to: {xlsx}")
